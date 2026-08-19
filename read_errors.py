import openpyxl

wb = openpyxl.load_workbook("DermaSense_Web_Test_Report.xlsx")
ws = wb.active

for row in range(4, 24):
    tc_id = ws.cell(row, 1).value
    status = ws.cell(row, 4).value
    if status == "[FAIL]":
        actual = ws.cell(row, 5).value
        error = ws.cell(row, 6).value
        print(f"{tc_id}: {actual}")
        print(f"Error: {error}\n")
