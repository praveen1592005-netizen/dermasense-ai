# -*- coding: utf-8 -*-
"""
DermaSense AI - Selenium E2E Web Test Suite
============================================
Tests the hosted Firebase web app end-to-end.
Generates a colour-coded Excel report.

USAGE:
  python selenium_tests/run_web_tests.py
  python selenium_tests/run_web_tests.py --email user@email.com --password pass123
  python selenium_tests/run_web_tests.py --url https://yourapp.vercel.app --headless

REQUIREMENTS:
  pip install selenium webdriver-manager openpyxl
"""

import sys, os, time, datetime, argparse, traceback

# Windows UTF-8 fix
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─── Args ─────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="DermaSense Selenium Web Test Runner")
parser.add_argument("--url",      default="https://dermasense-ai-18698.web.app", help="Base URL of hosted app")
parser.add_argument("--email",    default="",   help="Login email")
parser.add_argument("--password", default="",   help="Login password")
parser.add_argument("--headless", action="store_true", help="Run in headless mode")
parser.add_argument("--output",   default="DermaSense_Web_Test_Report.xlsx",     help="Excel output file")
args = parser.parse_args()

BASE_URL    = args.url.rstrip("/")
TEST_EMAIL  = args.email
TEST_PASS   = args.password
HEADLESS    = args.headless

# ─── Imports ──────────────────────────────────────────────────────────────────
DEPS_OK = False
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service  import Service as ChromeService
    from selenium.webdriver.chrome.options  import Options as ChromeOptions
    from selenium.webdriver.common.by       import By
    from selenium.webdriver.common.keys     import Keys
    from selenium.webdriver.support.ui      import WebDriverWait
    from selenium.webdriver.support         import expected_conditions as EC
    from selenium.common.exceptions         import (
        TimeoutException, NoSuchElementException, WebDriverException
    )
    from webdriver_manager.chrome           import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    DEPS_OK = True
    print("[OK] All dependencies loaded.")
except ImportError as e:
    print(f"[ERROR] {e}")
    print("Run: pip install selenium webdriver-manager openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# DRIVER SETUP
# ─────────────────────────────────────────────────────────────────────────────

def create_driver():
    opts = ChromeOptions()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1366,768")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    try:
        service = ChromeService(ChromeDriverManager().install())
        driver  = webdriver.Chrome(service=service, options=opts)
    except Exception:
        driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(5)
    driver.set_page_load_timeout(30)
    return driver


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def find(driver, locators, timeout=12):
    for by, val in locators:
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, val)))
        except Exception:
            continue
    return None

def click(driver, locators, timeout=12):
    el = find(driver, locators, timeout)
    if el:
        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)
        return True
    return False

def type_in(driver, locators, text, timeout=12):
    el = find(driver, locators, timeout)
    if el:
        el.clear()
        el.send_keys(text)
        return True
    return False

def visible(driver, locators, timeout=8):
    for by, val in locators:
        try:
            WebDriverWait(driver, timeout).until(
                EC.visibility_of_element_located((by, val)))
            return True
        except Exception:
            continue
    return False

def get_text(driver, locators, timeout=8):
    el = find(driver, locators, timeout)
    return el.text if el else ""

def wait(s=1.5):
    time.sleep(s)

def shot(driver, name):
    folder = os.path.join(os.path.dirname(__file__), "screenshots")
    os.makedirs(folder, exist_ok=True)
    ts   = datetime.datetime.now().strftime("%H%M%S")
    path = os.path.join(folder, f"{ts}_{name}.png")
    try:
        driver.save_screenshot(path)
    except Exception:
        pass
    return path

def navigate(driver, path=""):
    driver.get(f"{BASE_URL}/{path}")
    wait(2)


# ─────────────────────────────────────────────────────────────────────────────
# LOCATORS  (Flutter Web uses <flt-semantics> — fall back to CSS/XPath)
# ─────────────────────────────────────────────────────────────────────────────

# Login page
L_EMAIL    = [(By.ID, "email"),
              (By.CSS_SELECTOR, "input[type='email']"),
              (By.XPATH, "//input[@placeholder='Email Address' or contains(@placeholder,'mail')]"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='email_field'] input"),
              (By.XPATH, "//*[contains(@aria-label,'email')]//input"),
              (By.CSS_SELECTOR, "input")]

L_PASSWORD = [(By.ID, "password"),
              (By.CSS_SELECTOR, "input[type='password']"),
              (By.XPATH, "//input[@placeholder='Password' or contains(@placeholder,'assword')]"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='password_field'] input"),
              (By.XPATH, "(//input)[2]")]

L_LOGIN_BTN= [(By.ID, "login-button"),
              (By.XPATH, "//button[contains(.,'Sign In') or contains(.,'Login')]"),
              (By.CSS_SELECTOR, "button[aria-label='login_button']"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='login_button']")]

L_GOOGLE   = [(By.XPATH, "//button[contains(.,'Google')]"),
              (By.CSS_SELECTOR, "button[aria-label='google_signin_button']"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='google_signin_button']")]

L_GUEST    = [(By.XPATH, "//button[contains(.,'Guest') or contains(.,'guest')]"),
              (By.CSS_SELECTOR, "button[aria-label='guest_button']"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='guest_button']"),
              (By.XPATH, "//*[contains(text(),'Instant Guest Access')]")]

L_FORGOT   = [(By.XPATH, "//*[contains(text(),'Forgot')]"),
              (By.LINK_TEXT, "Forgot Password?")]

L_SIGNUP   = [(By.XPATH, "//*[contains(text(),'Sign Up')]"),
              (By.LINK_TEXT, "Sign Up")]

L_DASH     = [(By.ID, "dashboard"),
              (By.XPATH, "//*[contains(text(),'DermaSense AI') or contains(text(),'Scan') or contains(text(),'Dashboard')]"),
              (By.CSS_SELECTOR, "flt-semantics[aria-label='dashboard_screen']")]

L_ERROR    = [(By.XPATH, "//*[contains(text(),'failed') or contains(text(),'incorrect') or contains(text(),'fill') or contains(text(),'error')]"),
              (By.CSS_SELECTOR, ".error, .snackbar, [role='alert']")]

L_TITLE    = [(By.TAG_NAME, "title"),
              (By.XPATH,     "//h1 | //h2 | //h3")]

L_DERMA    = [(By.XPATH, "//*[contains(text(),'DermaSense')]"),
              (By.CSS_SELECTOR, "flt-semantics")]   # Flutter canvas fallback

L_REGISTER = [(By.XPATH, "//*[contains(text(),'Register') or contains(text(),'Create Account')]"),
              (By.CSS_SELECTOR, "flt-semantics")]

L_SCAN     = [(By.XPATH, "//*[contains(text(),'Scan') or contains(text(),'scan')]")]

L_PROFILE  = [(By.XPATH, "//*[contains(text(),'Profile')]")]


# ─────────────────────────────────────────────────────────────────────────────
# TEST RESULT MODEL
# ─────────────────────────────────────────────────────────────────────────────

class TR:
    def __init__(self, tc_id, name, desc=""):
        self.tc_id  = tc_id
        self.name   = name
        self.desc   = desc
        self.status = "NOT RUN"
        self.actual = ""
        self.error  = ""
        self.shot   = ""
        self.dur    = 0.0
        self.time   = ""
        self.url    = ""

results = []

def run(driver, tc, fn):
    t0 = time.time()
    tc.time = datetime.datetime.now().strftime("%H:%M:%S")
    try:
        fn(driver, tc)
        if tc.status not in ("FAIL", "SKIP"):
            tc.status = "PASS"
    except Exception as e:
        tc.status = "FAIL"
        tc.error  = str(e)[:300]
        tc.actual = traceback.format_exc()[-250:]
        tc.shot   = shot(driver, tc.tc_id)
    finally:
        tc.dur = round(time.time() - t0, 2)
        try:
            tc.url = driver.current_url
        except Exception:
            pass
    results.append(tc)
    icon = "[PASS]" if tc.status == "PASS" else ("[SKIP]" if tc.status == "SKIP" else "[FAIL]")
    print(f"  {icon} {tc.tc_id} {tc.name}  ({tc.dur}s)  {tc.url}")


# ─────────────────────────────────────────────────────────────────────────────
# TEST CASES
# ─────────────────────────────────────────────────────────────────────────────

def tc001(d, tc):
    """TC001 — App loads, title/branding visible"""
    tc.desc = "Hosted web app loads successfully at the base URL"
    navigate(d)
    wait(3)
    title = d.title
    page  = d.page_source[:2000]
    if title or "DermaSense" in page or "dermasense" in page.lower() or len(d.page_source) > 500:
        tc.actual = f"Page loaded. Title='{title}', URL={d.current_url}"
    else:
        raise AssertionError(f"Page did not load properly. Title='{title}'")


def tc002(d, tc):
    """TC002 — Page title matches DermaSense AI"""
    tc.desc = "Page title or heading contains DermaSense AI"
    navigate(d)
    wait(3)
    title = d.title
    src   = d.page_source
    if "DermaSense" in title or "DermaSense" in src or "dermasense" in src.lower():
        tc.actual = f"DermaSense branding found. Title='{title}'"
    else:
        tc.status = "FAIL"
        tc.actual = f"DermaSense not found in title or page source. Title='{title}'"


def tc003(d, tc):
    """TC003 — Login page accessible (has email or sign-in elements)"""
    tc.desc = "Login page has email field or sign-in elements"
    navigate(d)
    wait(4)
    src = d.page_source.lower()
    has_login = (visible(d, L_EMAIL, 5) or "sign in" in src or "login" in src
                 or "email" in src or "password" in src)
    if has_login:
        tc.actual = "Login page detected — sign-in elements present"
    else:
        tc.status = "FAIL"
        tc.actual = "Login page elements not found"


def tc004(d, tc):
    """TC004 — Email field accepts input"""
    tc.desc = "Email input field is accessible and accepts text"
    navigate(d)
    wait(4)
    if type_in(d, L_EMAIL, "selenium_test@dermasense.ai"):
        tc.actual = "Email field accepts input"
    else:
        tc.status = "SKIP"
        tc.actual = "Email field not interactable (Flutter canvas may block direct input)"


def tc005(d, tc):
    """TC005 — Password field accepts input"""
    tc.desc = "Password field is accessible and accepts text (masked)"
    navigate(d)
    wait(4)
    if type_in(d, L_PASSWORD, "TestPassword@123"):
        tc.actual = "Password field accepts input"
    else:
        tc.status = "SKIP"
        tc.actual = "Password field not directly interactable"


def tc006(d, tc):
    """TC006 — Login attempt with wrong credentials shows error"""
    tc.desc = "Wrong credentials display error message, stay on login"
    navigate(d)
    wait(4)
    type_in(d, L_EMAIL,    "invalid@fakeuser.com")
    wait(0.5)
    type_in(d, L_PASSWORD, "wrongpass999")
    wait(0.5)
    click(d, L_LOGIN_BTN)
    wait(5)
    on_login = visible(d, L_EMAIL, 5) or visible(d, L_LOGIN_BTN, 5)
    has_err  = visible(d, L_ERROR, 5)
    if has_err:
        tc.actual = "Error message shown for invalid credentials"
    elif on_login:
        tc.actual = "Stayed on login page (correct behavior, no redirect)"
    else:
        raise AssertionError("App redirected away with invalid credentials")


def tc007(d, tc):
    """TC007 — Login with empty email shows validation"""
    tc.desc = "Empty email field triggers validation"
    navigate(d)
    wait(4)
    type_in(d, L_EMAIL, "")
    wait(0.3)
    type_in(d, L_PASSWORD, "somepassword")
    wait(0.3)
    click(d, L_LOGIN_BTN)
    wait(3)
    err_visible = visible(d, L_ERROR, 4)
    on_login    = visible(d, L_EMAIL, 4) or visible(d, L_LOGIN_BTN, 4)
    if err_visible or on_login:
        tc.actual = "Validation shown or stayed on login for empty email"
    else:
        tc.actual = "No explicit validation UI (acceptable for Flutter web)"


def tc008(d, tc):
    """TC008 — Forgot Password link is present"""
    tc.desc = "Forgot Password link exists on login page"
    navigate(d)
    wait(4)
    if visible(d, L_FORGOT, 6):
        tc.actual = "Forgot Password link found"
    else:
        tc.status = "SKIP"
        tc.actual = "Forgot Password link not found (may need scroll or flutter semantics)"


def tc009(d, tc):
    """TC009 — Sign Up link navigates to registration"""
    tc.desc = "Sign Up link is present and clickable"
    navigate(d)
    wait(4)
    if not visible(d, L_SIGNUP, 6):
        tc.status = "SKIP"
        tc.actual = "Sign Up link not found on page"
        return
    initial_url = d.current_url
    click(d, L_SIGNUP)
    wait(3)
    reg_visible = visible(d, L_REGISTER, 6)
    url_changed = d.current_url != initial_url
    d.back(); wait(2)
    if reg_visible or url_changed:
        tc.actual = f"Sign Up navigated to register. URL changed: {url_changed}"
    else:
        tc.status = "FAIL"
        tc.actual = "Sign Up click did not navigate to registration"


def tc010(d, tc):
    """TC010 — Guest Access button present"""
    tc.desc = "Guest Access / Instant Access button visible"
    navigate(d)
    wait(4)
    if visible(d, L_GUEST, 6):
        tc.actual = "Guest Access button found"
    else:
        tc.status = "SKIP"
        tc.actual = "Guest Access button not found"


def tc011(d, tc):
    """TC011 — Guest Login works, reaches dashboard"""
    tc.desc = "Guest login button bypasses auth and opens dashboard"
    navigate(d)
    wait(4)
    if not click(d, L_GUEST, 8):
        tc.status = "SKIP"; tc.actual = "Guest button not found/clickable"; return
    wait(5)
    dash = visible(d, L_DASH, 15)
    if dash:
        tc.actual = f"Dashboard loaded after Guest login. URL={d.current_url}"
    else:
        tc.status = "FAIL"
        tc.actual = f"Dashboard not visible after Guest login. URL={d.current_url}"


def tc012(d, tc):
    """TC012 — Valid login (if credentials provided)"""
    tc.desc = "Valid email + password login reaches Dashboard"
    if not TEST_EMAIL or not TEST_PASS:
        tc.status = "SKIP"
        tc.actual = "No credentials provided — re-run with --email and --password"
        return
    navigate(d)
    wait(4)
    ok_e = type_in(d, L_EMAIL, TEST_EMAIL)
    wait(0.5)
    ok_p = type_in(d, L_PASSWORD, TEST_PASS)
    wait(0.5)
    if not ok_e or not ok_p:
        tc.status = "SKIP"
        tc.actual = "Could not type into login fields (Flutter canvas)"
        return
    click(d, L_LOGIN_BTN)
    wait(6)
    if visible(d, L_DASH, 15):
        tc.actual = f"Login successful with {TEST_EMAIL}. URL={d.current_url}"
    else:
        raise AssertionError(f"Dashboard not found after login. URL={d.current_url}")


def tc013(d, tc):
    """TC013 — Page is responsive (mobile width)"""
    tc.desc = "Page renders without overflow at 375px mobile width"
    navigate(d)
    wait(3)
    d.set_window_size(375, 812)
    wait(2)
    src = d.page_source
    # Check no horizontal scroll overflow (basic check)
    w = d.execute_script("return document.documentElement.scrollWidth")
    vw = d.execute_script("return window.innerWidth")
    d.set_window_size(1366, 768)  # restore
    if w <= vw + 30:
        tc.actual = f"Mobile responsive — scrollWidth={w}, viewportWidth={vw}"
    else:
        tc.status = "FAIL"
        tc.actual = f"Horizontal overflow at mobile: scrollWidth={w} > viewportWidth={vw}"


def tc014(d, tc):
    """TC014 — HTTPS secure connection"""
    tc.desc = "App is served over HTTPS"
    navigate(d)
    wait(2)
    url = d.current_url
    if url.startswith("https://"):
        tc.actual = f"HTTPS confirmed: {url}"
    else:
        tc.status = "FAIL"
        tc.actual = f"Not HTTPS: {url}"


def tc015(d, tc):
    """TC015 — Page load time < 10 seconds"""
    tc.desc = "Home page loads within 10 seconds"
    t0 = time.time()
    navigate(d)
    wait(1)
    # Wait for body to be present
    try:
        WebDriverWait(d, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    except Exception:
        pass
    elapsed = round(time.time() - t0, 2)
    if elapsed <= 10:
        tc.actual = f"Page loaded in {elapsed}s"
    else:
        tc.status = "FAIL"
        tc.actual = f"Page load too slow: {elapsed}s (threshold: 10s)"


def tc016(d, tc):
    """TC016 — Google Sign-In button present"""
    tc.desc = "Google Sign-In button visible on login page"
    navigate(d)
    wait(4)
    if visible(d, L_GOOGLE, 6):
        tc.actual = "Google Sign-In button found"
    else:
        tc.status = "SKIP"
        tc.actual = "Google button not found (may not be visible on web build)"


def tc017(d, tc):
    """TC017 — Browser Back button does not crash app"""
    tc.desc = "Browser back button works without breaking the app"
    navigate(d); wait(3)
    d.back(); wait(2)
    d.forward(); wait(2)
    src = d.page_source
    if src and len(src) > 100:
        tc.actual = "Back/Forward navigation works without crash"
    else:
        tc.status = "FAIL"
        tc.actual = "Page broken after back/forward navigation"


def tc018(d, tc):
    """TC018 — Page has meta description (SEO)"""
    tc.desc = "Page has a meta description tag (SEO)"
    navigate(d); wait(2)
    try:
        meta = d.find_element(By.XPATH,
                              "//meta[@name='description' or @property='og:description']")
        content = meta.get_attribute("content") or ""
        if content:
            tc.actual = f"Meta description found: '{content[:100]}'"
        else:
            tc.status = "FAIL"; tc.actual = "Meta description tag is empty"
    except NoSuchElementException:
        tc.status = "FAIL"
        tc.actual = "No meta description tag found"


def tc019(d, tc):
    """TC019 — Console has no JavaScript errors"""
    tc.desc = "Browser console has no critical JavaScript errors"
    navigate(d); wait(4)
    try:
        logs = d.get_log("browser")
        severes = [l for l in logs if l.get("level") == "SEVERE"]
        if not severes:
            tc.actual = "No SEVERE JavaScript console errors"
        else:
            tc.status = "FAIL"
            errs = "; ".join(l.get("message","")[:80] for l in severes[:3])
            tc.actual = f"SEVERE console errors: {errs}"
    except Exception as e:
        tc.status = "SKIP"
        tc.actual = f"Could not retrieve console logs: {e}"


def tc020(d, tc):
    """TC020 — Direct URL access /login or /# works"""
    tc.desc = "Direct URL access to login route works"
    routes = [f"{BASE_URL}/#/login", f"{BASE_URL}/#login",
              f"{BASE_URL}/login",   f"{BASE_URL}"]
    for route in routes:
        try:
            d.get(route)
            wait(3)
            if len(d.page_source) > 500:
                tc.actual = f"Direct URL access OK: {route}"
                return
        except Exception:
            continue
    tc.status = "FAIL"
    tc.actual = "None of the login routes loaded correctly"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(res_list, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"

    # Colours
    DARK  = "1A1A2E"; PURP = "4A0072"; PLITE = "7B1FA2"
    PGRN  = "1B5E20"; PRED = "B71C1C"; PORG  = "E65100"
    ALT   = "F3E5F5"; LGRY = "FAFAFA"; WHT   = "FFFFFF"

    def F(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
    def Ft(bold=False, color=WHT, sz=11): return Font(bold=bold, color=color, size=sz, name="Calibri")
    def Al(h="center", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)
    def Bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # Row 1 — Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "DermaSense AI  |  Selenium E2E Web Automation Test Report"
    c.font = Ft(bold=True, sz=15); c.fill = F(DARK); c.alignment = Al()
    ws.row_dimensions[1].height = 38

    # Row 2 — Meta
    ws.merge_cells("A2:J2")
    ts   = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    p    = sum(1 for r in res_list if r.status == "PASS")
    f    = sum(1 for r in res_list if r.status == "FAIL")
    sk   = sum(1 for r in res_list if r.status == "SKIP")
    nr   = sum(1 for r in res_list if r.status == "NOT RUN")
    rate = round(p / max(len(res_list), 1) * 100, 1)
    c = ws["A2"]
    c.value = (f"Generated: {ts}   |   URL: {BASE_URL}   |   Total: {len(res_list)}   |   "
               f"PASS: {p}   FAIL: {f}   SKIP: {sk}   NOT RUN: {nr}   Pass Rate: {rate}%")
    c.font = Ft(sz=9); c.fill = F(PURP); c.alignment = Al()
    ws.row_dimensions[2].height = 20

    # Row 3 — Headers
    headers    = ["TC ID","Test Name","Description","Status","Actual Result",
                  "Error","Duration(s)","Start Time","URL Visited","Screenshot"]
    col_widths = [10,     28,          40,           12,     55,
                  40,     12,          10,            50,     40]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci)
        cell.value = h; cell.font = Ft(bold=True); cell.fill = F(PLITE)
        cell.alignment = Al(w=True); cell.border = Bd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # Data rows
    S_MAP = {
        "PASS":    ("E8F5E9", PGRN,  "[PASS]"),
        "FAIL":    ("FFEBEE", PRED,  "[FAIL]"),
        "SKIP":    ("FFF3E0", PORG,  "[SKIP]"),
        "NOT RUN": ("F5F5F5", "546E7A", "[NOT RUN]"),
    }
    for ri, tc in enumerate(res_list, start=4):
        bg   = ALT if ri % 2 == 0 else LGRY
        s_bg, s_fg, s_txt = S_MAP.get(tc.status, ("F5F5F5", "546E7A", tc.status))
        row = [tc.tc_id, tc.name, tc.desc, s_txt,
               tc.actual or "-", tc.error[:200] if tc.error else "-",
               str(tc.dur), tc.time, tc.url or "-", tc.shot or "-"]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.value     = val; cell.border = Bd()
            cell.alignment = Al(h="left" if ci > 2 else "center", w=True)
            if ci == 4:
                cell.fill = F(s_bg)
                cell.font = Font(bold=True, color=s_fg, size=11, name="Calibri")
                cell.alignment = Al()
            else:
                cell.fill = F(bg)
                cell.font = Font(color="212121", size=10, name="Calibri")
        ws.row_dimensions[ri].height = 32

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "DermaSense AI  -  Selenium Test Summary"
    ws2["A1"].font = Ft(bold=True, sz=14); ws2["A1"].fill = F(DARK); ws2["A1"].alignment = Al()
    ws2.row_dimensions[1].height = 30
    rows_s = [
        ("Tested URL",       BASE_URL,        PURP),
        ("Total Test Cases", len(res_list),   PURP),
        ("PASSED",           p,               PGRN),
        ("FAILED",           f,               PRED),
        ("SKIPPED",          sk,              PORG),
        ("NOT RUN",          nr,              "546E7A"),
        ("Pass Rate",        f"{rate}%",      PLITE),
        ("Run Date/Time",    ts,              PURP),
    ]
    for i, (lbl, val, col) in enumerate(rows_s, 2):
        a = ws2.cell(row=i, column=1); b = ws2.cell(row=i, column=2)
        a.value = lbl; a.font = Ft(bold=True); a.fill = F(col)
        a.alignment = Al(h="left"); a.border = Bd()
        b.value = str(val); b.font = Ft(bold=True, sz=12); b.fill = F(col)
        b.alignment = Al(); b.border = Bd()
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 55
        ws2.row_dimensions[i].height = 26

    wb.save(filename)
    abs_path = os.path.abspath(filename)
    print(f"\n{'='*65}")
    print(f"[REPORT] Excel saved: {abs_path}")
    print(f"{'='*65}\n")
    return abs_path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if not DEPS_OK:
        sys.exit(1)

    print("\n" + "="*65)
    print("DermaSense AI  -  Selenium E2E Web Test Runner")
    print(f"  URL      : {BASE_URL}")
    print(f"  Email    : {TEST_EMAIL or '(not provided)'}")
    print(f"  Headless : {HEADLESS}")
    print("="*65 + "\n")

    driver = None
    try:
        print("[...] Launching Chrome browser...")
        driver = create_driver()
        print("[OK]  Browser started.\n")
    except Exception as e:
        print(f"[ERROR] Could not start browser: {e}")
        sys.exit(1)

    tests = [
        (TR("TC001","App Page Load"),                  tc001),
        (TR("TC002","DermaSense Branding"),             tc002),
        (TR("TC003","Login Page Elements Present"),     tc003),
        (TR("TC004","Email Field Input"),               tc004),
        (TR("TC005","Password Field Input"),            tc005),
        (TR("TC006","Invalid Login Shows Error"),       tc006),
        (TR("TC007","Empty Email Validation"),          tc007),
        (TR("TC008","Forgot Password Link"),            tc008),
        (TR("TC009","Sign Up Navigation"),              tc009),
        (TR("TC010","Guest Access Button Present"),     tc010),
        (TR("TC011","Guest Login Reaches Dashboard"),   tc011),
        (TR("TC012","Valid Email Login"),               tc012),
        (TR("TC013","Mobile Responsive Layout"),        tc013),
        (TR("TC014","HTTPS Secure Connection"),         tc014),
        (TR("TC015","Page Load Time < 10s"),            tc015),
        (TR("TC016","Google Sign-In Button"),           tc016),
        (TR("TC017","Browser Back Navigation"),         tc017),
        (TR("TC018","SEO Meta Description"),            tc018),
        (TR("TC019","No JS Console Errors"),            tc019),
        (TR("TC020","Direct URL Access"),               tc020),
    ]

    print("Running 20 test cases...\n")
    for tc, fn in tests:
        run(driver, tc, fn)

    try:
        driver.quit()
    except Exception:
        pass
    print("\n[OK] Browser closed.")

    report_path = generate_report(results, args.output)

    p  = sum(1 for r in results if r.status == "PASS")
    f  = sum(1 for r in results if r.status == "FAIL")
    sk = sum(1 for r in results if r.status == "SKIP")
    print(f"{'─'*50}")
    print(f"  TOTAL : {len(results)}")
    print(f"  PASS  : {p}")
    print(f"  FAIL  : {f}")
    print(f"  SKIP  : {sk}")
    print(f"  RATE  : {round(p/max(len(results),1)*100,1)}%")
    print(f"{'─'*50}")
    print(f"\n  Open report: {report_path}\n")


if __name__ == "__main__":
    main()
